import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, Alignment, PatternFill

# f = open('stex_.txt', 'w')  # открытие в режиме записи
# print('Введите количество стадий')
# Nr = int(input())
# print('Введите количество веществ')
# Nv = int(input())
# for v in range(Nv):
#     f.write('Y' + str(v + 1) + ' ')
# f.write('Obr')
# f.write('\n')
# stex_matrix = np.zeros((Nr, Nv), dtype=int)
# for i in range(Nr):
#     print(i + 1, '- я стадия')
#     print('Введите количество веществ, участвующих в стадии')
#     R_V = int(input())
#     print('Далее вводите номера веществ, соответствующие их порядку и их значения')
#     number, value = [], []
#     for j in range(R_V):
#         numb = int(input())
#         number.append(numb)
#         val = int(input())
#         value.append(val)
#     print('Хотите повторить? y/n')
#     print(number)
#     y = input()
#     while y != 'n':
#         number, value = [], []
#         numb = int(input())
#         number.append(numb)
#         val = int(input())
#         value.append(val)
#         print('Хотите повторить? y/n')
#         y = input()
#     for z in range(R_V):
#         stex_matrix[i][number[z] - 1] = value[z]
#     for k in range(Nv):
#         f.write(str(stex_matrix[i][k]) + ' ')
#     print('Обратимость')
#     obr = int(input())
#     f.write(str(obr))
#     f.write('\n')

wb = load_workbook(filename=r'C:\Users\Acer\Desktop\Магистратура\Магистерская диссертация\Python\каталитический риформинг бензина\stex_matrix.xlsx')
ws = wb.active
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=col).value != 0:
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor="FF0000")
wb.save("stex_koef.xlsx")